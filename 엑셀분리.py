# https://www.youtube.com/watch?v=qwKcEVjFew0
# 파이썬 + 엑셀 = openpyxl (파이썬으로 엑셀 다루기)
import time
import openpyxl

print("START ... ", time.strftime('%y.%m.%d - %X')) # 년월일 - 시간

target_file = "D:\Documents\My Passport\쇼핑몰\hblee666_오너클랜상품리스트(230128-1W)_PLAYAUTO_1661592_1_1.xlsx"
wb = openpyxl.load_workbook(target_file, data_only=True)

ws = wb['PLAYAUTO']  # 시트선택
#ws = wb.active      # 시트를 읽어온다


########## 저장할 엑셀 파일
step = 1000

s_path = "D:/Documents/My Passport/쇼핑몰/"
s_file = "to_" + str(step) + ".xlsx"
s_wb = openpyxl.Workbook()

# 시트 삭제
#s_wb.remove(s_ws)
s_ws = s_wb.create_sheet('PLAYAUTO')
#s_ws.cell(row = 1, column = 1).value = '1,1'

row_cnt = 0
row1_value = []

# 타이틀을 보관해 둔다
for row1 in ws.iter_rows(min_row=1, max_row=1):
   for cell in row1:
      row1_value.append(cell.value)
s_ws.append(row1_value)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row) :
   # 선택한 워크시트를 한줄씩 읽어서 저장 한다
   
   row_cnt = row_cnt + 1
   row_value = []
   
   if ( row == '' ) :
      break
   
   for cell in row:
      row_value.append(cell.value)
   
   s_ws.append(row_value)

   if ((row_cnt % 1000) == 0):         # 1000건씩 다른 파일로 저장한다
      s_wb.remove(s_wb['Sheet'])  # 자동 생성되는 시트는 지운다
      s_wb['PLAYAUTO']            # 시트선택
      s_wb.save(s_path + s_file)
      print("#1 ==========================", s_file, step, row_cnt)
      
      # 새로운 work book을 만든다
      s_wb = openpyxl.Workbook()
      s_ws = s_wb.create_sheet('PLAYAUTO')
      step = step + 1000
      s_file = "to_"+ str(step) + ".xlsx"
      s_ws.append( row1_value )  # 타이틀은 먼저 SET 한다 
   
s_wb.remove(s_wb['Sheet'])  # 자동 생성되는 시트는 지운다
s_wb['PLAYAUTO']            # 시트선택
s_wb.save(s_path + s_file)
print("#9 ==========================", s_file, step, row_cnt)
   
print("END ... ", time.strftime('%y.%m.%d - %X')) # 년월일 - 시간
   
   
   